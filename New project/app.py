from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT_DIR = SCRIPT_DIR.parent
WORKBOOK_FILENAME = "Self_Employed_Shift_Manager.xlsx"
LOCAL_DOWNLOADS_WORKBOOK = Path(r"C:/Users/olugb/Downloads/Self_Employed_Shift_Manager.xlsx")


def detect_default_workbook_path() -> Path:
    candidates = [
        SCRIPT_DIR / WORKBOOK_FILENAME,
        REPO_ROOT_DIR / WORKBOOK_FILENAME,
        Path.cwd() / WORKBOOK_FILENAME,
        LOCAL_DOWNLOADS_WORKBOOK,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return SCRIPT_DIR / WORKBOOK_FILENAME


DEFAULT_WORKBOOK_PATH = detect_default_workbook_path()
MAX_SHIFT_ROWS = 500

SHIFT_COLUMNS = [
    "Date",
    "Month",
    "Client",
    "Shift Type",
    "Hours Worked",
    "Base Rate (GBP)",
    "Overtime Hours",
    "Overtime Rate (GBP)",
    "Mileage (Miles)",
    "Mileage Rate (GBP)",
    "Travel (GBP)",
    "Food (GBP)",
    "Other (GBP)",
]

NUMERIC_COLUMNS = [
    "Hours Worked",
    "Base Rate (GBP)",
    "Overtime Hours",
    "Overtime Rate (GBP)",
    "Mileage (Miles)",
    "Mileage Rate (GBP)",
    "Travel (GBP)",
    "Food (GBP)",
    "Other (GBP)",
]

MULTIPLIER_BY_SHIFT_TYPE = {
    "Standard": 1.0,
    "Weekend": 1.25,
    "Night": 1.5,
    "Bank Holiday": 2.0,
}

MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def normalize_shifts(df: pd.DataFrame) -> pd.DataFrame:
    for column in SHIFT_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA

    df = df[SHIFT_COLUMNS].copy()

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Month"] = df["Date"].dt.strftime("%b")
    df["Shift Type"] = df["Shift Type"].fillna("Standard").replace("", "Standard")
    df["Client"] = df["Client"].fillna("").astype(str).str.strip()

    for column in NUMERIC_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0.0)

    multiplier = df["Shift Type"].map(MULTIPLIER_BY_SHIFT_TYPE).fillna(1.0)
    df["Gross Pay (GBP)"] = (df["Hours Worked"] * df["Base Rate (GBP)"] * multiplier) + (
        df["Overtime Hours"] * df["Overtime Rate (GBP)"]
    )
    df["Total Expenses (GBP)"] = (
        (df["Mileage (Miles)"] * df["Mileage Rate (GBP)"])
        + df["Travel (GBP)"]
        + df["Food (GBP)"]
        + df["Other (GBP)"]
    )
    df["Net Pay (GBP)"] = df["Gross Pay (GBP)"] - df["Total Expenses (GBP)"]

    has_data = (
        df["Date"].notna()
        | df["Client"].ne("")
        | df["Hours Worked"].ne(0)
        | df["Base Rate (GBP)"].ne(0)
        | df["Overtime Hours"].ne(0)
        | df["Overtime Rate (GBP)"].ne(0)
        | df["Mileage (Miles)"].ne(0)
        | df["Mileage Rate (GBP)"].ne(0)
        | df["Travel (GBP)"].ne(0)
        | df["Food (GBP)"].ne(0)
        | df["Other (GBP)"].ne(0)
    )

    return df.loc[has_data].reset_index(drop=True)


def to_python_date(value: Any) -> date | None:
    if pd.isna(value):
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def shifts_from_workbook(workbook_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=False)
    if "Shift Records" not in wb.sheetnames:
        return pd.DataFrame(columns=SHIFT_COLUMNS)

    ws = wb["Shift Records"]
    rows: list[dict[str, Any]] = []
    for row_index in range(2, MAX_SHIFT_ROWS + 2):
        values = [ws[f"{col}{row_index}"].value for col in "ABCDEFGHIJKLM"]
        if all(value is None or value == "" for value in values):
            continue

        rows.append(
            {
                "Date": values[0],
                "Month": values[1],
                "Client": values[2],
                "Shift Type": values[3],
                "Hours Worked": values[4],
                "Base Rate (GBP)": values[5],
                "Overtime Hours": values[6],
                "Overtime Rate (GBP)": values[7],
                "Mileage (Miles)": values[8],
                "Mileage Rate (GBP)": values[9],
                "Travel (GBP)": values[10],
                "Food (GBP)": values[11],
                "Other (GBP)": values[12],
            }
        )

    if not rows:
        return pd.DataFrame(columns=SHIFT_COLUMNS)
    return pd.DataFrame(rows)


def workbook_with_updates(template_bytes: bytes, shifts: pd.DataFrame) -> bytes:
    wb = load_workbook(BytesIO(template_bytes), data_only=False)
    if "Shift Records" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'Shift Records' sheet.")

    ws = wb["Shift Records"]

    for row_index in range(2, MAX_SHIFT_ROWS + 2):
        for col in "ABCDEFGHIJKLM":
            ws[f"{col}{row_index}"].value = None

    cleaned = normalize_shifts(shifts).head(MAX_SHIFT_ROWS).copy()

    for idx, row in cleaned.iterrows():
        row_index = idx + 2
        ws[f"A{row_index}"] = to_python_date(row["Date"])
        ws[f"B{row_index}"] = row["Month"] if pd.notna(row["Month"]) else None
        ws[f"C{row_index}"] = row["Client"] if row["Client"] else None
        ws[f"D{row_index}"] = row["Shift Type"] if pd.notna(row["Shift Type"]) else "Standard"
        ws[f"E{row_index}"] = float(row["Hours Worked"])
        ws[f"F{row_index}"] = float(row["Base Rate (GBP)"])
        ws[f"G{row_index}"] = float(row["Overtime Hours"])
        ws[f"H{row_index}"] = float(row["Overtime Rate (GBP)"])
        ws[f"I{row_index}"] = float(row["Mileage (Miles)"])
        ws[f"J{row_index}"] = float(row["Mileage Rate (GBP)"])
        ws[f"K{row_index}"] = float(row["Travel (GBP)"])
        ws[f"L{row_index}"] = float(row["Food (GBP)"])
        ws[f"M{row_index}"] = float(row["Other (GBP)"])

    if "Client Summary" in wb.sheetnames:
        ws_client = wb["Client Summary"]
        for row_index in range(2, 102):
            ws_client[f"A{row_index}"] = None
        clients = [name for name in cleaned["Client"].dropna().astype(str).str.strip().unique() if name]
        for idx, client in enumerate(clients[:100], start=2):
            ws_client[f"A{idx}"] = client

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def monthly_summary(shifts: pd.DataFrame) -> pd.DataFrame:
    if shifts.empty:
        return pd.DataFrame(
            {
                "Month": MONTH_ORDER,
                "Total Gross (GBP)": [0.0] * 12,
                "Total Expenses (GBP)": [0.0] * 12,
                "Net Income (GBP)": [0.0] * 12,
            }
        )

    grouped = (
        shifts.groupby("Month", dropna=False)[["Gross Pay (GBP)", "Total Expenses (GBP)", "Net Pay (GBP)"]]
        .sum()
        .rename(
            columns={
                "Gross Pay (GBP)": "Total Gross (GBP)",
                "Total Expenses (GBP)": "Total Expenses (GBP)",
                "Net Pay (GBP)": "Net Income (GBP)",
            }
        )
    )
    grouped = grouped.reindex(MONTH_ORDER, fill_value=0.0).reset_index().rename(columns={"index": "Month"})
    return grouped


def client_summary(shifts: pd.DataFrame) -> pd.DataFrame:
    if shifts.empty:
        return pd.DataFrame(columns=["Client Name", "Total Gross (GBP)", "Total Net (GBP)"])

    filtered = shifts[shifts["Client"].astype(str).str.strip().ne("")]
    if filtered.empty:
        return pd.DataFrame(columns=["Client Name", "Total Gross (GBP)", "Total Net (GBP)"])

    grouped = (
        filtered.groupby("Client")[["Gross Pay (GBP)", "Net Pay (GBP)"]]
        .sum()
        .reset_index()
        .rename(
            columns={
                "Client": "Client Name",
                "Gross Pay (GBP)": "Total Gross (GBP)",
                "Net Pay (GBP)": "Total Net (GBP)",
            }
        )
        .sort_values("Total Net (GBP)", ascending=False)
    )
    return grouped


def tax_estimate(shifts: pd.DataFrame) -> pd.DataFrame:
    total_net = float(shifts["Net Pay (GBP)"].sum()) if not shifts.empty else 0.0
    estimated_tax = total_net * 0.20
    estimated_ni = total_net * 0.09
    pension = total_net * 0.05
    take_home = total_net - estimated_tax - estimated_ni - pension

    return pd.DataFrame(
        {
            "Metric": [
                "Total Net Income (GBP)",
                "Estimated Tax (20%)",
                "Estimated National Insurance (9%)",
                "Suggested Pension Saving (5%)",
                "Estimated Take Home (GBP)",
            ],
            "Amount (GBP)": [total_net, estimated_tax, estimated_ni, pension, take_home],
        }
    )


def load_source_bytes(uploaded_file: Any, workbook_path: str) -> tuple[bytes | None, str]:
    if uploaded_file is not None:
        return uploaded_file.getvalue(), f"upload::{uploaded_file.name}::{uploaded_file.size}"

    raw_path = (workbook_path or "").strip()
    if not raw_path:
        search_paths = [DEFAULT_WORKBOOK_PATH]
    else:
        path = Path(raw_path).expanduser()
        if path.is_absolute():
            search_paths = [path]
        else:
            search_paths = [
                Path.cwd() / path,
                SCRIPT_DIR / path,
                REPO_ROOT_DIR / path,
            ]

    seen: set[Path] = set()
    for candidate in search_paths:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists():
            return resolved.read_bytes(), f"path::{resolved}::{resolved.stat().st_mtime}"

    return None, "none"


def currency_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    formatted = df.copy()
    for column in columns:
        if column in formatted.columns:
            numeric = pd.to_numeric(formatted[column], errors="coerce").fillna(0.0)
            formatted[column] = numeric.map(lambda value: f"£{value:,.2f}")
    return formatted


def main() -> None:
    import streamlit as st

    st.set_page_config(page_title="Self-Employed Shift Manager", layout="wide")
    st.title("Self-Employed Shift Manager")
    st.caption("Application built from your Excel workbook structure.")

    with st.sidebar:
        st.subheader("Workbook Source")
        workbook_path = st.text_input("Excel file path", str(DEFAULT_WORKBOOK_PATH))
        uploaded_file = st.file_uploader("Or upload workbook", type=["xlsx"])

    source_bytes, source_key = load_source_bytes(uploaded_file, workbook_path)

    if source_bytes is None:
        st.error("Workbook not found. Upload a .xlsx file or update the file path in the sidebar.")
        return

    if st.session_state.get("source_key") != source_key:
        st.session_state["source_key"] = source_key
        st.session_state["template_bytes"] = source_bytes
        loaded = shifts_from_workbook(source_bytes)
        st.session_state["shifts"] = normalize_shifts(loaded)

    shifts_df = st.session_state.get("shifts", pd.DataFrame(columns=SHIFT_COLUMNS))

    st.subheader("Shift Records")
    editable = shifts_df.reindex(columns=SHIFT_COLUMNS, fill_value=pd.NA)

    edited = st.data_editor(
        editable,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Date": st.column_config.DateColumn("Date", format="YYYY-MM-DD"),
            "Month": st.column_config.TextColumn("Month", disabled=True),
            "Client": st.column_config.TextColumn("Client"),
            "Shift Type": st.column_config.SelectboxColumn(
                "Shift Type",
                options=["Standard", "Weekend", "Night", "Bank Holiday"],
            ),
            "Hours Worked": st.column_config.NumberColumn("Hours Worked", min_value=0.0, step=0.5, format="%.2f"),
            "Base Rate (GBP)": st.column_config.NumberColumn(
                "Base Rate (GBP)", min_value=0.0, step=0.5, format="%.2f"
            ),
            "Overtime Hours": st.column_config.NumberColumn(
                "Overtime Hours", min_value=0.0, step=0.5, format="%.2f"
            ),
            "Overtime Rate (GBP)": st.column_config.NumberColumn(
                "Overtime Rate (GBP)", min_value=0.0, step=0.5, format="%.2f"
            ),
            "Mileage (Miles)": st.column_config.NumberColumn(
                "Mileage (Miles)", min_value=0.0, step=1.0, format="%.2f"
            ),
            "Mileage Rate (GBP)": st.column_config.NumberColumn(
                "Mileage Rate (GBP)", min_value=0.0, step=0.01, format="%.2f"
            ),
            "Travel (GBP)": st.column_config.NumberColumn("Travel (GBP)", min_value=0.0, step=0.5, format="%.2f"),
            "Food (GBP)": st.column_config.NumberColumn("Food (GBP)", min_value=0.0, step=0.5, format="%.2f"),
            "Other (GBP)": st.column_config.NumberColumn("Other (GBP)", min_value=0.0, step=0.5, format="%.2f"),
        },
    )

    normalized = normalize_shifts(edited)
    st.session_state["shifts"] = normalized

    total_gross = float(normalized["Gross Pay (GBP)"].sum()) if not normalized.empty else 0.0
    total_expenses = float(normalized["Total Expenses (GBP)"].sum()) if not normalized.empty else 0.0
    total_net = float(normalized["Net Pay (GBP)"].sum()) if not normalized.empty else 0.0

    metric_1, metric_2, metric_3 = st.columns(3)
    metric_1.metric("Total Gross", f"£{total_gross:,.2f}")
    metric_2.metric("Total Expenses", f"£{total_expenses:,.2f}")
    metric_3.metric("Total Net", f"£{total_net:,.2f}")

    st.dataframe(
        currency_columns(
            normalized[
                SHIFT_COLUMNS
                + [
                    "Gross Pay (GBP)",
                    "Total Expenses (GBP)",
                    "Net Pay (GBP)",
                ]
            ],
            [
                "Base Rate (GBP)",
                "Overtime Rate (GBP)",
                "Mileage Rate (GBP)",
                "Travel (GBP)",
                "Food (GBP)",
                "Other (GBP)",
                "Gross Pay (GBP)",
                "Total Expenses (GBP)",
                "Net Pay (GBP)",
            ],
        ),
        use_container_width=True,
        hide_index=True,
    )

    tab_month, tab_client, tab_tax = st.tabs(["Monthly Summary", "Client Summary", "Tax Estimate"])

    with tab_month:
        month_df = monthly_summary(normalized)
        st.dataframe(
            currency_columns(month_df, ["Total Gross (GBP)", "Total Expenses (GBP)", "Net Income (GBP)"]),
            use_container_width=True,
            hide_index=True,
        )
        st.bar_chart(month_df.set_index("Month")[["Net Income (GBP)"]], use_container_width=True)

    with tab_client:
        client_df = client_summary(normalized)
        st.dataframe(
            currency_columns(client_df, ["Total Gross (GBP)", "Total Net (GBP)"]),
            use_container_width=True,
            hide_index=True,
        )

    with tab_tax:
        tax_df = tax_estimate(normalized)
        st.dataframe(currency_columns(tax_df, ["Amount (GBP)"]), use_container_width=True, hide_index=True)

    try:
        updated_workbook = workbook_with_updates(st.session_state["template_bytes"], normalized)
        st.download_button(
            "Download Updated Workbook",
            data=updated_workbook,
            file_name="Self_Employed_Shift_Manager_updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as exc:
        st.warning(f"Could not generate updated workbook: {exc}")


if __name__ == "__main__":
    main()

